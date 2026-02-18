import os
import re
import shutil
import pdfplumber
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askdirectory
from datetime import datetime
from openpyxl import load_workbook

# ------------------------------
# Select Folder
# ------------------------------
Tk().withdraw()
folder_path = askdirectory(title="Select Folder Containing Uber PDF Bills")

if not folder_path:
    print("No folder selected.")
    exit()

output_folder = os.path.join(folder_path, "Refined")
os.makedirs(output_folder, exist_ok=True)

# ------------------------------
# Extraction Function
# ------------------------------

def extract_details(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() + "\n"

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    # ---------------- DATE ----------------
    ride_date = None
    date_match = re.search(r'([A-Za-z]+ \d{1,2}, \d{4})', text)

    if date_match:
        date_str = date_match.group(1)
        for fmt in ("%b %d, %Y", "%B %d, %Y"):
            try:
                ride_date = datetime.strptime(date_str, fmt).date()
                break
            except:
                continue

    # ---------------- FARE ----------------
    total_match = re.search(r'Total ₹\s*([\d,]+\.\d{2})', text)
    fare_amount = float(total_match.group(1).replace(",", "")) if total_match else 0.0

    # ---------------- LICENSE PLATE ----------------
    license_plate = ""
    plate_pattern = re.compile(r'[A-Z]{2}[0-9]{2}[A-Z]{1,2}[0-9]{4}')

    for line in lines:
        cleaned = re.sub(r'[^A-Za-z0-9]', '', line).upper()
        match = plate_pattern.search(cleaned)
        if match:
            license_plate = match.group(0)
            break

    # ---------------- LOCATION EXTRACTION ----------------
    start_location = ""
    drop_location = ""

    # ----- FORMAT 2 (Pipe format: 6:26 pm | address) -----
    pipe_pattern = re.compile(r'\d{1,2}:\d{2}\s*(am|pm)\s*\|\s*(.+)', re.IGNORECASE)

    pipe_matches = []
    for line in lines:
        m = pipe_pattern.search(line)
        if m:
            pipe_matches.append(m.group(2).strip())

    if len(pipe_matches) >= 2:
        start_location = pipe_matches[0]
        drop_location = pipe_matches[1]

    else:
        # ----- FORMAT 1 (Multiline format after License Plate) -----
        time_pattern = re.compile(r'^\d{1,2}:\d{2}\s*(am|pm)$', re.IGNORECASE)

        license_index = None
        for i, line in enumerate(lines):
            if "License Plate" in line:
                license_index = i
                break

        if license_index is not None:
            time_indexes = []
            for i in range(license_index, len(lines)):
                if time_pattern.match(lines[i]):
                    time_indexes.append(i)

            if len(time_indexes) >= 2:

                # Start
                start_block = []
                i = time_indexes[0] + 1
                while i < len(lines) and not time_pattern.match(lines[i]):
                    start_block.append(lines[i])
                    i += 1
                start_location = " ".join(start_block)

                # Drop
                drop_block = []
                i = time_indexes[1] + 1
                while i < len(lines):
                    line = lines[i]

                    if (
                        time_pattern.match(line) or
                        "You rode with" in line or
                        "Want to review" in line or
                        "http" in line
                    ):
                        break

                    drop_block.append(line)
                    i += 1

                drop_location = " ".join(drop_block)

    return ride_date, fare_amount, start_location, drop_location, license_plate

# ------------------------------
# Process PDFs
# ------------------------------

records = []

for file in os.listdir(folder_path):
    if not file.lower().endswith(".pdf"):
        continue

    original_path = os.path.join(folder_path, file)

    try:
        ride_date, fare_amount, start, drop, plate = extract_details(original_path)

        if ride_date:
            new_filename = f"{ride_date.strftime('%Y%m%d')}_{fare_amount:.2f}.pdf"
        else:
            new_filename = f"UnknownDate_{fare_amount:.2f}.pdf"

        new_path = os.path.join(output_folder, new_filename)

        counter = 1
        base_name = new_filename
        while os.path.exists(new_path):
            new_filename = base_name.replace(".pdf", f"_{counter}.pdf")
            new_path = os.path.join(output_folder, new_filename)
            counter += 1

        shutil.copy2(original_path, new_path)

        records.append({
            "File Name": new_filename,
            "Date": ride_date,
            "Start Location": start,
            "Drop Location": drop,
            "CAB License Plate": plate,
            "Fare Amount": fare_amount,
            "Full Path": new_path
        })

    except Exception as e:
        print(f"Error processing {file}: {e}")

df = pd.DataFrame(records)

# ------------------------------
# Save Excel
# ------------------------------

excel_path = os.path.join(output_folder, "Uber_Bills_Summary.xlsx")

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df.drop(columns=["Full Path"]).to_excel(writer, index=False)

# ------------------------------
# Format Excel (Column Width + Date Format)
# ------------------------------

wb = load_workbook(excel_path)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter

    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    ws.column_dimensions[column].width = min(max_length + 3, 60)

# Date column formatting (Column B)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    for cell in row:
        cell.number_format = 'DD-MM-YYYY'

# Add hyperlinks
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=1).hyperlink = df.iloc[row-2]["Full Path"]
    ws.cell(row=row, column=1).style = "Hyperlink"

wb.save(excel_path)

print("\n✅ All Done – Both Formats Supported")
print(f"Refined folder created at:\n{output_folder}")
input("\nPress Enter to exit...")